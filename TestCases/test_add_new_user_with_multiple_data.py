from typing import Any

import json
import jsonpath
import requests
import openpyxl


def test_add_new_student():
    ## API_URL code
    api_url = "http://thetestingworldapi.com/api/studentsDetails"
    f = open("C:/Users/bhatp/Desktop/API-PYTEST/add_new_student.json",'r')
    json_requests = json.loads(f.read())
    ## Excel code
    wk = openpyxl.load_workbook("C:/Users/bhatp/Desktop/API-PYTEST/testdata.xlsx")
    sh = wk['Sheet1']
    rows = sh.max_row
    for i in range(2,rows+1):
        cell_first_name = sh.cell(row=i,column=1)
        cell_mid_name = sh.cell(row=i, column=2)
        cell_last_name = sh.cell(row=i, column=3)
        cell_dob = sh.cell(row=i, column=4)
        json_requests['first_name'] = cell_first_name.value
        json_requests['middle_name'] = cell_mid_name.value
        json_requests['last_name'] = cell_last_name.value
        json_requests['date_of_birth'] = cell_dob.value
        response = requests.post(api_url, json_requests)
        print(response.text)
        print(response.status_code)
        assert response.status_code == 201


